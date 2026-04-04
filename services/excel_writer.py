"""
Excel writing services.

Writes reconciliation outputs to an Excel workbook using openpyxl (via pandas
ExcelWriter or direct openpyxl usage for formatting).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class ExcelWriteOptions:
    """
    Options controlling output workbook generation.

    This will later include formatting options, sheet naming, freeze panes, etc.
    """

    output_path: Path


def write_reconciliation_output(
    *,
    purchase_cleaned_df: pd.DataFrame,
    gstr_b2b_cleaned_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    detailed_reco_df: pd.DataFrame,
    options: ExcelWriteOptions,
) -> None:
    """
    Write the GST reconciliation output workbook with professional formatting.

    Args:
        purchase_cleaned_df: Cleaned purchase register DataFrame.
        gstr_b2b_cleaned_df: Cleaned GSTR-2B B2B DataFrame.
        summary_df: GSTIN-level reconciliation summary DataFrame.
        detailed_reco_df: Invoice-level reconciliation DataFrame with `status`.
        options: ExcelWriteOptions controlling output path and formatting.
    """
    options.output_path.parent.mkdir(parents=True, exist_ok=True)

    sheets: dict[str, pd.DataFrame] = {
        "Purchase Cleaned": purchase_cleaned_df,
        "GSTR B2B Cleaned": gstr_b2b_cleaned_df,
        "Reconciliation Summary": summary_df,
        "Detailed Reconciliation": detailed_reco_df,
    }

    with pd.ExcelWriter(options.output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        workbook = writer.book
        for sheet_name in sheets:
            ws = workbook[sheet_name]
            _apply_basic_formatting(ws)

        _highlight_detailed_reconciliation_rows(workbook["Detailed Reconciliation"])


def write_workbook(*, sheets: dict[str, pd.DataFrame], options: ExcelWriteOptions) -> None:
    """
    Generic workbook writer retained for reusable ad-hoc exports.
    """
    options.output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(options.output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)


def _apply_basic_formatting(ws: Worksheet) -> None:
    """
    Apply base professional formatting to a worksheet.
    """
    if ws.max_row < 1 or ws.max_column < 1:
        return

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Header row styling.
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Body styling.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize_columns(ws)


def _autosize_columns(ws: Worksheet) -> None:
    """
    Auto-fit column widths with safe limits.
    """
    min_width = 10
    max_width = 45
    pad = 2

    for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        adjusted = min(max(max_len + pad, min_width), max_width)
        ws.column_dimensions[col_cells[0].column_letter].width = adjusted


def _highlight_detailed_reconciliation_rows(ws: Worksheet) -> None:
    """
    Highlight rows in Detailed Reconciliation based on status value.

    The underlying cell values are not changed; only cell styles are updated.
    """
    status_col_index = _find_column_index(ws, "status")
    if status_col_index is None or ws.max_row < 2:
        return

    fills = {
        "NOT IN 2B": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # red-ish
        "NOT IN BOOKS": PatternFill(
            start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
        ),  # orange-ish
        "AMOUNT MISMATCH": PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        ),  # yellow
    }

    for row_idx in range(2, ws.max_row + 1):
        status_value = ws.cell(row=row_idx, column=status_col_index).value
        if status_value is None:
            continue
        status = str(status_value).strip().upper()
        fill = fills.get(status)
        if fill is None:
            continue
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _find_column_index(ws: Worksheet, column_name: str) -> int | None:
    """
    Find a header column index by name (case-insensitive).
    """
    target = column_name.strip().lower()
    for idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=idx).value
        if value is not None and str(value).strip().lower() == target:
            return idx
    return None

